import json
import os
import re
import pandas as pd
from z3 import *
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

def load_json_file(file_path):
    """加载JSON文件"""
    with open(file_path, 'r', encoding='utf-8') as f:
        return json.load(f)

def split_condition(condition):
    """
    将条件分解为原子条件
    
    Args:
        condition (str): 条件表达式
    
    Returns:
        list: 原子条件列表
    """
    # 处理复杂条件的情况
    # 首先尝试使用正则表达式匹配基本的比较表达式
    basic_expressions = re.findall(r'[^&|()]+?(?:==|!=|>=|<=|>|<)[^&|()]+', condition)
    
    # 如果没有找到基本表达式，则返回整个条件作为一个原子
    if not basic_expressions:
        return [condition]
    
    # 清理原子条件
    atoms = [atom.strip() for atom in basic_expressions if atom.strip()]
    
    return atoms

def evaluate_atom_condition(atom, test_case, constants):
    """
    评估原子条件的真假
    
    Args:
        atom (str): 原子条件
        test_case (str): 测试用例
        constants (dict): 常量字典
    
    Returns:
        bool: 条件是否为真
    """
    # 解析测试用例中的变量赋值
    variables = {}
    for assignment in test_case.split(';'):
        assignment = assignment.strip()
        if not assignment:
            continue
        
        parts = assignment.split('=')
        if len(parts) == 2:
            var_name = parts[0].strip()
            var_value = parts[1].strip()
            
            # 处理布尔值
            if var_value.lower() == 'true':
                var_value = 1
            elif var_value.lower() == 'false':
                var_value = 0
            else:
                # 尝试将值转换为数字
                try:
                    var_value = float(var_value)
                    # 如果是整数，转换为整数
                    if var_value.is_integer():
                        var_value = int(var_value)
                except ValueError:
                    pass
            
            variables[var_name] = var_value
    
    # 合并变量和常量
    all_vars = {**variables, **constants}
    
    # 替换原子条件中的变量
    modified_atom = atom
    
    # 替换true和false为1和0
    modified_atom = re.sub(r'\btrue\b', '1', modified_atom, flags=re.IGNORECASE)
    modified_atom = re.sub(r'\bfalse\b', '0', modified_atom, flags=re.IGNORECASE)
    
    # 按照变量名长度排序，避免部分替换问题（例如，先替换N2R25再替换N2R25Design）
    sorted_vars = sorted(all_vars.keys(), key=len, reverse=True)
    
    for var_name in sorted_vars:
        var_value = all_vars[var_name]
        # 确保只替换完整的变量名，而不是变量名的一部分
        pattern = r'\b' + re.escape(var_name) + r'\b'
        modified_atom = re.sub(pattern, str(var_value), modified_atom)
    
    # 清理修改后的原子条件，确保它是一个有效的表达式
    # 移除多余的括号
    while modified_atom.startswith('(') and modified_atom.endswith(')'):
        # 检查括号是否匹配
        if modified_atom.count('(') == modified_atom.count(')'):
            modified_atom = modified_atom[1:-1]
        else:
            break
    
    # 尝试使用Python的eval评估条件
    try:
        # 创建一个安全的局部变量环境
        local_vars = {}
        # 使用eval评估条件
        result = eval(modified_atom, {"__builtins__": {}}, local_vars)
        return bool(result)
    except Exception as e:
        print(f"无法评估条件 '{atom}' (修改后: '{modified_atom}'): {e}")
        # 如果无法评估，默认为真
        return True

def modify_condition_based_on_evaluation(condition, test_case, constants):
    """
    基于评估结果修改条件
    
    Args:
        condition (str): 原始条件
        test_case (str): 测试用例
        constants (dict): 常量字典
    
    Returns:
        str: 修改后的条件
    """
    # 使用正则表达式直接评估整个条件
    # 解析测试用例中的变量赋值
    variables = {}
    for assignment in test_case.split(';'):
        assignment = assignment.strip()
        if not assignment:
            continue
        
        parts = assignment.split('=')
        if len(parts) == 2:
            var_name = parts[0].strip()
            var_value = parts[1].strip()
            
            # 处理布尔值
            if var_value.lower() == 'true':
                var_value = 1
            elif var_value.lower() == 'false':
                var_value = 0
            else:
                # 尝试将值转换为数字
                try:
                    var_value = float(var_value)
                    # 如果是整数，转换为整数
                    if var_value.is_integer():
                        var_value = int(var_value)
                except ValueError:
                    pass
            
            variables[var_name] = var_value
    
    # 合并变量和常量
    all_vars = {**variables, **constants}
    
    # 替换条件中的变量
    modified_condition = condition
    
    # 替换true和false为1和0
    modified_condition = re.sub(r'\btrue\b', '1', modified_condition, flags=re.IGNORECASE)
    modified_condition = re.sub(r'\bfalse\b', '0', modified_condition, flags=re.IGNORECASE)
    
    # 按照变量名长度排序，避免部分替换问题
    sorted_vars = sorted(all_vars.keys(), key=len, reverse=True)
    
    for var_name in sorted_vars:
        var_value = all_vars[var_name]
        # 确保只替换完整的变量名，而不是变量名的一部分
        pattern = r'\b' + re.escape(var_name) + r'\b'
        modified_condition = re.sub(pattern, str(var_value), modified_condition)
    
    # 尝试评估整个条件
    try:
        # 创建一个安全的局部变量环境
        local_vars = {}
        # 使用eval评估条件
        result = eval(modified_condition, {"__builtins__": {}}, local_vars)
        
        # 如果条件为假，则在整个条件外添加!()
        if not result:
            return f"!({condition})"
        else:
            return condition
    except Exception as e:
        print(f"无法评估整个条件 '{condition}' (修改后: '{modified_condition}'): {e}")
        
        # 如果无法评估整个条件，则尝试分解为原子条件
        atoms = split_condition(condition)
        modified_condition = condition
        
        for atom in atoms:
            is_true = evaluate_atom_condition(atom, test_case, constants)
            if not is_true:
                # 如果原子条件为假，在条件外添加!()
                # 替换原子条件为其否定形式
                modified_condition = modified_condition.replace(atom, f"!({atom})")
        
        return modified_condition

def generate_excel(testcases_file, constants_file, output_file):
    """
    生成测试用例Excel文件
    
    Args:
        testcases_file (str): 测试用例JSON文件路径
        constants_file (str): 常量JSON文件路径
        output_file (str): 输出Excel文件路径
    """
    # 加载测试用例和常量
    testcases = load_json_file(testcases_file)
    constants = load_json_file(constants_file)
    
    # 准备Excel数据
    data = []
    
    # 遍历测试用例
    for testcase in testcases:
        req_id = testcase.get("requirement_id", "")
        module_name = testcase.get("module_name", "")
        precondition = testcase.get("precondition", "")
        condition = testcase.get("condition", "")
        true_result = testcase.get("true_result", "")
        false_result = testcase.get("false_result", "")
        
        # 处理true_test_case - 条件保持不变
        for test in testcase.get("true_test_case", []):
            data.append([
                req_id,
                module_name,
                precondition,
                condition,
                test,
                true_result
            ])
        
        # 处理false_test_case - 条件取反
        for test in testcase.get("false_test_case", []):
            # 对于false_test_case，直接在整个条件外添加!()
            modified_condition = f"!({condition})"
            
            data.append([
                req_id,
                module_name,
                precondition,
                modified_condition,
                test,
                false_result
            ])
    
    # 创建DataFrame
    df = pd.DataFrame(data, columns=[
        "需求编号", "模块名称", "前置条件", "判断条件", "测试用例", "结果"
    ])
    
    # 保存为Excel
    df.to_excel(output_file, index=False)
    
    # 使用openpyxl美化Excel
    wb = openpyxl.load_workbook(output_file)
    ws = wb.active
    
    # 设置列宽
    ws.column_dimensions['A'].width = 15  # 需求编号
    ws.column_dimensions['B'].width = 30  # 模块名称
    ws.column_dimensions['C'].width = 15  # 前置条件
    ws.column_dimensions['D'].width = 40  # 判断条件
    ws.column_dimensions['E'].width = 40  # 测试用例
    ws.column_dimensions['F'].width = 40  # 结果
    
    # 设置表头样式
    header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    header_font = Font(bold=True)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # 设置单元格样式
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical='center', wrap_text=True)
    
    # 保存美化后的Excel
    wb.save(output_file)
    
    print(f"测试用例已生成至: {output_file}")

if __name__ == "__main__":
    # 确保输出目录存在
    os.makedirs("output", exist_ok=True)
    
    # 文件路径
    testcases_file = "output/testcase_parsed_merged_with_results.json"
    constants_file = "output/data_constants.json"
    output_file = "output/generated_testcases.xlsx"
    
    # 生成Excel
    generate_excel(testcases_file, constants_file, output_file)
